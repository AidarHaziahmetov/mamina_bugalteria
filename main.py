from openpyxl import load_workbook
import os
for i in range(len(os.listdir())):
    print(i, os.listdir()[i]) 
def generate_new_id_card(old_id_card: str) -> str:
    return (10-len(old_id_card))*'0'+old_id_card


data_file = load_workbook(f'{os.listdir()[int(input("Введие номер файла который нужно переделать в отчет: "))]}')
list0 = list(data_file[data_file.sheetnames[0]].values)
start = 0
end = len(list0)
summa = 0
nd = []
for i in range(len(list0)):
    list0[i] = list(list0[i])
    for j in list0[i]:
        if type(j) == str:
            if j.lower() == "номер карты":
                start = i+1
            elif j.lower() == "итого":
                end = i
for i in range(start, end):
    if type(list0[i][-1]) in (int,float):
        summa += list0[i][-1]
        list0[i][-1] = str(float(list0[i][-1]))
        if list0[i][-1][-2]=='.':
            list0[i][-1] += '0'
        list0[i][-1] = list0[i][-1].replace('.',',')

        if type(list0[i][0]) in (int,float):
            list0[i][0] = generate_new_id_card(str(list0[i][0]))
        else:
            list0[i][0] = generate_new_id_card(str(input("Введите отсутствующий ID(в файле он на строке под номером "+str(i+1)+'): ')))
        nd.append([list0[i][0],list0[i][-1]])
summa = str(round(float(summa),2))
if summa[-2]=='.':
            summa += '0'
summa = summa.replace('.',',')

def date_generator(string: str) -> str:
    numbers = '0123456789'
    for i in range(len(string)-1,0,-1):
        if string[i] in numbers:
            new_date_string = string[i-9:i+1]
            date_range = '01' + new_date_string[2:] + ' по ' + new_date_string
            return [new_date_string, date_range]

def shapka_generator(string: str) -> str:
    return f'Отчет по питанию Скания\nНазвание ресторана: Центральный офис, Столовая Scania ДЦ Уфа\nПериод с {string}\nТорговое предприятие\nУчетный день;Время открытия;Номер карты клиента;Столовая г.Уфа, ул. А. Локотченко, д. 2 \n'


date_array = date_generator(list0[2][0])

with open(input('Введите название будующего файла: ')+'.txt','w') as f:
    f.write(shapka_generator(date_array[1]))
    for i in nd:
        f.write(f';{date_array[0]} 12:45;{i[0]};{i[1]};;\n')
    f.write(f'{date_array[0]} всего;;;{summa};;')




